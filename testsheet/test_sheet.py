import ast
import os
import openpyxl
import logging
from typing import List, Dict, Optional
from openpyxl.styles import Font, PatternFill, Alignment
from enum import Enum


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class TestCaseTag(Enum):
    DESCRIPTION = "Description:"
    TEST_STEP = "Step:"
    EXPECTED_OUTPUT = "Expected Output:"
    PRECONDITION = "Precondition:"


class TestSheet:
    """
    A class to extract test cases from a test script file and export them to an Excel file.
    """

    def __init__(self, file_path: str, output_excel: str):
        """
        Initialize the TestSheet.

        Args:
            file_path (str): Path to the test script file or directory containing test cases.
            output_excel (str): Path to the output Excel file.
        """
        self.file_path = file_path
        self.output_excel = output_excel
        self.lines = []
        self.test_cases = []

    def read_file(self, file_path: str) -> None:
        """
        Read the contents of the input test script file.

        Args:
            file_path (str): Path to the test script file.

        Raises:
            FileNotFoundError: If the input file does not exist.
        """
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                self.lines = file.readlines()
        except FileNotFoundError:
            logging.error(f"File not found: {file_path}")
            raise

    def extract_function_info(self, node: ast.FunctionDef) -> Dict[str, str | None ]:
        """
        Extract information (description, pre-condition, steps, expected output) for a test function.

        Args:
            node (ast.FunctionDef): The AST node representing the test function.

        Returns:
            Dict[str, str]: A dictionary containing the extracted test case information.
        """
        description = self.extract_description(node)
        pre_condition = self.extract_pre_condition(node)
        test_steps = self.extract_test_steps(node)
        expected_output = self.extract_expected_output(node)

        test_case = {
            "Test Case Name": node.name,
            "Automation Status": "",
            "Test Case Description": description or "",
            "Pre Condition": "\n".join(pre_condition) if pre_condition else "",
            "Test Steps": "\n".join(test_steps) if test_steps else "",
            "Expected Output": "\n".join(expected_output) if expected_output else "",
            "Test Result": ""
        }

        return test_case

    def extract_description(self, node: ast.FunctionDef) -> Optional[str]:
        """
        Extract the description from comments above the function or decorators.

        Args:
            node (ast.FunctionDef): The AST node representing the test function.

        Returns:
            Optional[str]: The extracted description, or None if not found.
        """
        try:
            function_start = node.lineno - 1
            description_lines = []

            for i in range(function_start - 1, -1, -1):
                line = self.lines[i].strip()

                # Stop if we encounter a non-comment line that is not a decorator
                if not (line.startswith("#") or line.startswith("@") or line.startswith('"""') or line.startswith(
                        "'''")):
                    break

                # Handle multiline comments
                if line.startswith('"""') or line.startswith("'''"):
                    comment_content = []
                    i -= 1
                    while i >= 0:
                        line = self.lines[i].strip()
                        step_text = (
                            line.split(f'""" {TestCaseTag.DESCRIPTION.value}', 1)[
                                -1] if f'""" {TestCaseTag.DESCRIPTION.value}' in line else
                            line.split(f"''' {TestCaseTag.DESCRIPTION.value}", 1)[
                                -1] if f"''' {TestCaseTag.DESCRIPTION.value}" in line else
                            line
                        )

                        if line.startswith('"""') or line.startswith("'''"):
                            comment_content.append(step_text)
                            break
                        comment_content.append(step_text)
                        i -= 1

                    comment_content.reverse()
                    description_lines.extend(comment_content)
                    break

                # Handle single-line comments
                if line.startswith("#"):
                    if f"{TestCaseTag.DESCRIPTION.value}" in line:
                        description_lines.append(line.split(f"{TestCaseTag.DESCRIPTION.value}", 1)[1].strip())
                    else:
                        description_lines.append(line.lstrip("#").strip())
            if not description_lines:
                logging.info(f"Description not found for function '{node.name}'."
                             f"Please check the function definition and tag.")

            return " ".join(description_lines).strip() if description_lines else None
        except Exception as e:
            logging.error(
                f"Error during description extraction for function '{node.name}': "
                f"{e}. Please check the function definition and parameters.")

    def extract_pre_condition(self, node: ast.FunctionDef) -> list[str] | None:
        """
        Extract the pre-condition from comments above the function.

        Args:
            node (ast.FunctionDef): The AST node representing the test function.

        Returns:
             List[str]: The extracted pre-condition, or None if not found.
        """
        try:
            pre_condition = []
            function_start = node.lineno - 1
            function_end = node.end_lineno if hasattr(node, "end_lineno") else function_start + 1

            step_counter = 1
            i = function_start + 1
            while i < function_end:
                line = self.lines[i].strip()

                # Handle single-line comments
                if line.startswith(f"# {TestCaseTag.PRECONDITION.value}"):
                    pre_condition.append(f"{step_counter}. " + line.split(f"# {TestCaseTag.PRECONDITION.value}", 1)[1].strip())
                    step_counter += 1

                # Handle multiline comments
                elif line.startswith(f'""" {TestCaseTag.PRECONDITION.value}') or line.startswith(f"''' {TestCaseTag.PRECONDITION.value}"):
                    comment_lines = []
                    i += 0
                    while i < function_end:
                        line = self.lines[i].strip()
                        step_text = (
                            line.split(f'""" {TestCaseTag.PRECONDITION.value}', 1)[
                                -1] if f'""" {TestCaseTag.PRECONDITION.value}' in line else
                            line.split(f"''' {TestCaseTag.PRECONDITION.value}", 1)[
                                -1] if f"''' {TestCaseTag.PRECONDITION.value}" in line else
                            line
                        )
                        if line.endswith('"""') or line.endswith("'''"):
                            break
                        if "\\n" in step_text:
                            parts = step_text.split("\\n")
                            # Add each part as a separate line
                            formatted_output = "\n".join(parts)
                            comment_lines.append(formatted_output)
                        else:
                            comment_lines.append(step_text)
                        i += 1

                    # Treat the entire multiline comment block as a single step
                    if comment_lines:
                        pre_condition.append(f"{step_counter}. " + " ".join(comment_lines).strip())
                        step_counter += 1

                i += 1

            return pre_condition
        except Exception as e:
            logging.error(
                f"Error during precondition extraction for function '{node.name}': "
                f"{e}. Please check the function definition and parameters.")

    def extract_test_steps(self, node: ast.FunctionDef) -> list[str] | None:
        """
        Extract test steps from comments inside the function, including multiline comments.
        Each comment block is treated as a single step.

        Args:
            node (ast.FunctionDef): The AST node representing the test function.

        Returns:
            List[str]: A list of test steps.
        """
        try:
            test_steps = []
            function_start = node.lineno - 1
            function_end = node.end_lineno if hasattr(node, "end_lineno") else function_start + 1

            step_counter = 1
            i = function_start + 1
            while i < function_end:
                line = self.lines[i].strip()

                # Handle single-line comments
                if line.startswith(f"# {TestCaseTag.TEST_STEP.value}"):
                    test_steps.append(f"{step_counter}. " + line.split(f"# {TestCaseTag.TEST_STEP.value}", 1)[1].strip())
                    step_counter += 1

                # Handle multiline comments
                elif line.startswith(f'""" {TestCaseTag.TEST_STEP.value}') or line.startswith(f"''' {TestCaseTag.TEST_STEP.value}"):
                    comment_lines = []
                    i += 0
                    while i < function_end:
                        line = self.lines[i].strip()
                        step_text = (
                            line.split(f'""" {TestCaseTag.TEST_STEP.value}', 1)[
                                -1] if f'""" {TestCaseTag.TEST_STEP.value}' in line else
                            line.split(f"''' {TestCaseTag.TEST_STEP.value}", 1)[
                                -1] if f"''' {TestCaseTag.TEST_STEP.value}" in line else
                            line
                        )
                        if line.endswith('"""') or line.endswith("'''"):
                            break
                        if "\\n" in step_text:
                            parts = step_text.split("\\n")
                            # Add each part as a separate line
                            formatted_output = "\n".join(parts)
                            comment_lines.append(formatted_output)
                        else:
                            comment_lines.append(step_text)
                        i += 1

                    # Treat the entire multiline comment block as a single step
                    if comment_lines:
                        test_steps.append(f"{step_counter}. " + " ".join(comment_lines).strip())
                        step_counter += 1

                i += 1
            if not test_steps:
                logging.info(f"Test Steps not found for function '{node.name}'."
                             f"Please check the function definition and parameters.")

            return test_steps
        except Exception as e:
            logging.error(
                f"Error during test step extraction for function '{node.name}': "
                f"{e}. Please check the function definition and parameters.")

    def extract_expected_output(self, node: ast.FunctionDef) -> list[str] | None:
        """
        Extract expected output from comments inside the function.

        Args:
            node (ast.FunctionDef): The AST node representing the test function.

        Returns:
            List[str]: A list of expected outputs.
        """
        try:
            expected_output = []
            function_start = node.lineno - 1
            function_end = node.end_lineno if hasattr(node, "end_lineno") else function_start + 1

            step_counter = 1
            i = function_start + 1
            while i < function_end:
                line = self.lines[i].strip()

                # Handle single-line comments
                if line.startswith(f"# {TestCaseTag.EXPECTED_OUTPUT.value}"):
                    expected_output.append(f"{step_counter}. " + line.split(f"# {TestCaseTag.EXPECTED_OUTPUT.value}", 1)[1].strip())
                    step_counter += 1

                # Handle multiline comments
                elif line.startswith(f'""" {TestCaseTag.EXPECTED_OUTPUT.value}') or line.startswith(f"''' {TestCaseTag.EXPECTED_OUTPUT.value}"):
                    comment_lines = []
                    i += 0
                    while i < function_end:
                        line = self.lines[i].strip()
                        step_text = (
                            line.split(f'""" {TestCaseTag.EXPECTED_OUTPUT.value}', 1)[
                                -1] if f'""" {TestCaseTag.EXPECTED_OUTPUT.value}' in line else
                            line.split(f"''' {TestCaseTag.EXPECTED_OUTPUT.value}", 1)[
                                -1] if f"''' {TestCaseTag.EXPECTED_OUTPUT.value}" in line else
                            line
                        )
                        if line.endswith('"""') or line.endswith("'''"):
                            break
                        if "\\n" in step_text:
                            parts = step_text.split("\\n")
                            # Add each part as a separate line
                            formatted_output = "\n".join(parts)
                            comment_lines.append(formatted_output)
                        else:
                            comment_lines.append(step_text)
                        i += 1

                    # Treat the entire multiline comment block as a single step
                    if comment_lines:
                        expected_output.append(f"{step_counter}. " + " ".join(comment_lines).strip())
                        step_counter += 1

                i += 1
            if not expected_output:
                logging.info(f"Expected Output' not found for function '{node.name}'."
                             f"Please check the function definition and parameters.")
            return expected_output
        except Exception as e:
            logging.error(
                f"Error during Expected Output extraction for function '{node.name}': "
                f"{e}. Please check the function definition and parameters.")

    def parse_test_cases(self) -> None:
        """
        Parse the test script file and extract all test cases.
        """
        tree = ast.parse("".join(self.lines))
        for node in ast.walk(tree):
            if isinstance(node, ast.FunctionDef) and node.name.startswith("test_"):
                test_case = self.extract_function_info(node)
                self.test_cases.append(test_case)

    def write_to_excel(self, wb: openpyxl.Workbook, sheet_name: str) -> None:
        """
        Write the extracted test cases to an Excel sheet with formatting.

        Args:
            wb (openpyxl.Workbook): The workbook to which the sheet will be added.
            sheet_name (str): The name of the sheet to be created.
        """
        try:
            ws = wb.create_sheet(title=sheet_name)

            # Define headers
            headers = ["Test Case Name", "Automation Status", "Test Case Description",
                       "Pre Condition", "Test Steps", "Expected Output", "Test Result"]
            ws.append(headers)

            # **Format Header (Bold + Grey Background)**
            bold_font = Font(bold=True)
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            for cell in ws[1]:  # Apply formatting to first row (header)
                cell.font = bold_font
                cell.fill = grey_fill

            # Write test cases
            for case in self.test_cases:
                ws.append([case[key] for key in case])

            # **Apply Text Wrapping to All Cells**
            wrap_text = Alignment(wrap_text=True)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = wrap_text

            # **Auto-adjust Column Width**
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 30

            logging.info(f"Test cases written to sheet: {sheet_name}")

        except Exception as e:
            logging.error(f"Error writing to Excel: {e}")

    def process_file(self, file_path: str, wb: openpyxl.Workbook) -> None:
        """
        Process a single file and add its test cases to the workbook.

        Args:
            file_path (str): Path to the test script file.
            wb (openpyxl.Workbook): The workbook to which the sheet will be added.
        """
        self.read_file(file_path)
        self.parse_test_cases()
        sheet_name = os.path.splitext(os.path.basename(file_path))[0]
        self.write_to_excel(wb, sheet_name)
        self.test_cases = []  # Reset test cases for the next file

    def process_directory(self, directory_path: str, wb: openpyxl.Workbook) -> None:
        """
        Recursively process all test script files in the directory and add their test cases to the workbook.

        Args:
            directory_path (str): Path to the directory containing test script files.
            wb (openpyxl.Workbook): The workbook to which the sheets will be added.
        """
        for root, _, files in os.walk(directory_path):
            if "site-packages" in root:
                continue
            for file in files:
                if file.startswith("test_") and file.endswith(".py"):
                    file_path = os.path.join(root, file)
                    self.process_file(file_path, wb)

    def run(self) -> None:
        """
        Execute the entire process: read file(s), parse test cases, and write to Excel.
        """
        wb = openpyxl.Workbook()
        if os.path.isfile(self.file_path):
            self.process_file(self.file_path, wb)
        elif os.path.isdir(self.file_path):
            self.process_directory(self.file_path, wb)
        else:
            logging.error(f"Invalid path: {self.file_path}")
            return

        # Remove the default sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        wb.save(self.output_excel)
        logging.info(f"All test cases written to {self.output_excel}")
